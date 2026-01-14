import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from odoo import models
from odoo.modules.module import get_module_resource

class AntuxVacacionesReport(models.AbstractModel):
    _name = 'antux.vacaciones.report'
    _description = 'Reporte Registro de Vacaciones'

    def build_mensual_excel(self, batch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vacaciones de Control'

        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # -------------------------------------------------
        # ANCHOS DE COLUMNAS
        # -------------------------------------------------
        widths = [5, 18, 20, 18, 18, 8, 14, 14, 16, 20]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # -------------------------------------------------
        # TITULO
        # -------------------------------------------------
        ws.merge_cells('A1:J1')
        ws['A1'] = 'REGISTRO ANUAL DE VACACIONES'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 30

        # -------------------------------------------------
        # LOGO
        # -------------------------------------------------
        logo_path = get_module_resource(
            'antux_data_hub', 'static', 'img', 'logo_mtess.png'
        )
        if logo_path:
            img = Image(logo_path)
            img.width = 220
            img.height = 120
            ws.add_image(img, 'A3')

        # -------------------------------------------------
        # DATOS EMPRESA (rótulos)
        # -------------------------------------------------
        ws['E3'] = 'Año:'
        ws['E4'] = 'Localidad:'
        ws['E5'] = 'Departamento:'
        ws['E6'] = 'Correo Electrónico:'
        ws['E7'] = 'Registro Patronal N°:'

        ws['I3'] = 'Razón Social:'
        ws['I4'] = 'Explotación:'
        ws['I5'] = 'Domicilio:'
        ws['I6'] = 'RUC:'
        ws['I7'] = 'Teléfono:'

        for cell in ['E3','E4','E5','E6','E7','I3','I4','I5','I6','I7']:
            ws[cell].font = bold

        # -------------------------------------------------
        # ENCABEZADOS DE TABLA
        # -------------------------------------------------
        start = 9

        ws.merge_cells(f'G{start}:H{start}')
        ws[f'G{start}'] = 'Duración de Vacaciones'
        ws[f'G{start}'].font = bold
        ws[f'G{start}'].alignment = center
        ws[f'G{start}'].border = thin

        headers = [
            'N°', 'Apellidos', 'Nombres',
            'Fecha de Entrada', 'Fecha de Salida',
            'Días', 'Desde', 'Hasta',
            'Remuneración', 'Observaciones'
        ]

        header_row = start + 1
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = bold
            cell.alignment = center
            cell.border = thin

        ws.freeze_panes = f'A{header_row+1}'

        # -------------------------------------------------
        # DATOS DE VACACIONES (ESTIRAR REGISTROS)
        # -------------------------------------------------
        current_row = header_row + 1
        counter = 1

        for l in batch.vacaciones_line_ids:
            ws.cell(current_row, 1, counter).border = thin
            ws.cell(current_row, 1).alignment = center

            ws.cell(current_row, 2, l.last_name or '').border = thin
            ws.cell(current_row, 3, l.first_name or '').border = thin

            ws.cell(current_row, 4,l.desde_vacaciones.strftime('%d/%m/%Y') if l.desde_vacaciones else '').border = thin

            ws.cell(current_row, 5,l.hasta_vacaciones.strftime('%d/%m/%Y') if l.hasta_vacaciones else '').border = thin

            ws.cell(current_row, 6, l.dias_vacaciones or '').border = thin
            ws.cell(current_row, 6).alignment = center

            ws.cell(
                current_row, 7,
                l.desde_vacaciones.strftime('%d/%m/%Y') if l.desde_vacaciones else ''
            ).border = thin

            ws.cell(
                current_row, 8,
                l.hasta_vacaciones.strftime('%d/%m/%Y') if l.hasta_vacaciones else ''
            ).border = thin

            ws.cell(current_row, 9, l.remuneracion_vacaciones or '').border = thin
            ws.cell(current_row, 10, l.observaciones_vacaciones or '').border = thin

            current_row += 1
            counter += 1

        # -------------------------------------------------
        # SALIDA
        # -------------------------------------------------
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream


    def build_anual_excel(self, batch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vacaciones - Anual'

        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')

        # -------------------------------------------------
        # CABECERAS
        # -------------------------------------------------
        headers = [
            'N°',
            'Apellidos',
            'Nombres',
            'Fecha de Entrada',
            'Fecha de Salida',
            'Días',
            'Desde',
            'Hasta',
            'Remuneración',
            'Observaciones',
        ]

        header_row = 1
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.font = bold
            cell.alignment = center

        ws.freeze_panes = f'A{header_row + 1}'

        # -------------------------------------------------
        # ANCHOS FIJOS (EVITA ERRORES CON MERGED CELLS)
        # -------------------------------------------------
        widths = [5, 18, 18, 16, 16, 8, 14, 14, 16, 25]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # -------------------------------------------------
        # DATOS (ESTIRAR)
        # -------------------------------------------------
        current_row = header_row + 1
        counter = 1

        for l in batch.vacaciones_line_ids:
            ws.cell(current_row, 1, counter).alignment = center

            ws.cell(current_row, 2, l.last_name or '').alignment = left
            ws.cell(current_row, 3, l.first_name or '').alignment = left

            ws.cell(
                current_row, 4,
                l.desde_vacaciones.strftime('%d/%m/%Y') if l.desde_vacaciones else ''
            ).alignment = center

            ws.cell(
                current_row, 5,
                l.hasta_vacaciones.strftime('%d/%m/%Y') if l.hasta_vacaciones else ''
            ).alignment = center

            ws.cell(current_row, 6, l.dias_vacaciones or '').alignment = center

            ws.cell(
                current_row, 7,
                l.desde_vacaciones.strftime('%d/%m/%Y') if l.desde_vacaciones else ''
            ).alignment = center

            ws.cell(
                current_row, 8,
                l.hasta_vacaciones.strftime('%d/%m/%Y') if l.hasta_vacaciones else ''
            ).alignment = center

            ws.cell(current_row, 9, l.remuneracion_vacaciones or '').alignment = center
            ws.cell(current_row, 10, l.observaciones_vacaciones or '').alignment = left

            current_row += 1
            counter += 1

        # -------------------------------------------------
        # SALIDA
        # -------------------------------------------------
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
