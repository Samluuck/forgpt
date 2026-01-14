import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from odoo import models
from odoo.modules.module import get_module_resource

class AntuxResumenGeneralReport(models.AbstractModel):
    _name = 'antux.resumen.general.report'
    _description = 'Resumen General de Personas Ocupadas'

    def _get_data_summary(self, batch):
        """
        Calcula los totales por categoría y género.
        Categorías:
        1) Supervisores o Jefes
        2) Empleados
        3) Obreros
        4) Menores
        """
        data = {
            'supervisores': {'M': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0},
                             'F': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0}},
            'empleados':    {'M': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0},
                             'F': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0}},
            'obreros':      {'M': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0},
                             'F': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0}},
            'menores':      {'M': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0},
                             'F': {'count': 0, 'hours': 0, 'salary': 0, 'entries': 0, 'exits': 0}},
        }

        for line in batch.line_ids:
            # Determinar Género
            gender = 'M'
            if line.sex and line.sex.upper() in ['F', 'FEMENINO', 'MUJER']:
                gender = 'F'
            
            # Determinar Categoría
            category = 'empleados' # Default
            
            job_title = (line.job_title or '').upper()
            worker_type = (line.worker_type or '').upper()
            
            # Lógica de clasificación
            is_minor = False
            if line.fecha_nacimiento:
                # Calculo de edad aproximado al fin de año (o periodo)
                # Asumimos año del periodo del lote
                year = batch.period_id.date_from.year
                born = line.fecha_nacimiento.year
                age = year - born
                if age < 18:
                    is_minor = True

            if is_minor:
                category = 'menores'
            elif any(x in job_title for x in ['SUPERVISOR', 'JEFE', 'GERENTE', 'DIRECTOR']):
                category = 'supervisores'
            elif 'OBRERO' in worker_type or any(x in job_title for x in ['OBRERO', 'OPERARIO']):
                category = 'obreros'
            
            # Acumular
            d = data[category][gender]
            d['count'] += 1
            
            # Horas: Asumimos 8 horas * dias trabajados si no hay campo horas especifico
            # En sueldos_jornales se usaba days_worked * 8
            hours = (line.days_worked or 0) * 8
            d['hours'] += hours
            
            # Salario: Usamos salary_total o salary_base
            salary = line.salary_total or line.salary_base or 0.0
            d['salary'] += salary
            
            # Entradas / Salidas en el periodo
            # Chequeamos si la fecha cae en el rango del periodo
            p_start = batch.period_id.date_from
            p_end = batch.period_id.date_to
            
            if line.entry_date and p_start <= line.entry_date <= p_end:
                d['entries'] += 1
            
            if line.exit_date and p_start <= line.exit_date <= p_end:
                d['exits'] += 1

        return data

    def build_control_excel(self, batch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Resumen General'
        
        # Estilos
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Logo y Encabezado Institucional (Similar a otros reportes)
        ws.merge_cells('C1:H1')
        ws['C1'] = 'RESUMEN GENERAL DE PERSONAS OCUPADAS'
        ws['C1'].font = Font(bold=True, size=14)
        ws['C1'].alignment = center
        
        # Logo
        logo_path = get_module_resource('antux_data_hub', 'static', 'img', 'logo_mtess.png')
        if logo_path:
            img = Image(logo_path)
            img.width = 200
            img.height = 100
            ws.add_image(img, 'A2')

        # Datos Empresa
        ws['D3'] = 'Año'; ws['E3'] = batch.period_id.date_from.year
        ws['D4'] = 'Localidad:'; ws['E4'] = 'ASUNCION (DISTRITO)'
        ws['D5'] = 'Departamento:'; ws['E5'] = 'CAPITAL'
        ws['D6'] = 'Página Web:'; ws['E6'] = '-'
        ws['D7'] = 'Correo Electrónico:'; ws['E7'] = batch.company_id.email
        ws['D8'] = 'REGISTRO PATRONAL Nº:'; ws['E8'] = batch.line_ids[0].patronal_number if batch.line_ids else ''

        ws['G3'] = 'Razón Social:'; ws['H3'] = batch.company_id.name
        ws['G4'] = 'Explotación:'; ws['H4'] = 'OTRAS ACTIVIDADES...'
        ws['G5'] = 'Domicilio:'; ws['H5'] = batch.company_id.street
        ws['G6'] = 'Nº Patronal IPS:'; ws['H6'] = batch.line_ids[0].patronal_number if batch.line_ids else ''
        ws['G7'] = 'RUC:'; ws['H7'] = batch.company_id.vat
        ws['G8'] = 'Telefono:'; ws['H8'] = batch.company_id.phone
        
        for r in range(3, 9):
            ws[f'D{r}'].font = bold
            ws[f'G{r}'].font = bold

        # Tabla Principal
        start_row = 10
        self._build_matrix_table(ws, batch, start_row)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def build_tabular_excel(self, batch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Resumen Tabular'
        
        # En tabular simple, empezamos directo con la tabla
        # Opcional: Poner un titulo simple
        ws['A1'] = 'RESUMEN GENERAL DE PERSONAS OCUPADAS'
        ws['A1'].font = Font(bold=True, size=12)
        
        start_row = 3
        self._build_matrix_table(ws, batch, start_row)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    def _build_matrix_table(self, ws, batch, start_row):
        # Estilos
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Fila Titulos Superiores
        ws.merge_cells(f'B{start_row}:I{start_row}')
        ws[f'B{start_row}'] = 'DISCRIMINACION DE PERSONAS OCUPADAS POR CATEGORIA'
        ws[f'B{start_row}'].border = border
        ws[f'B{start_row}'].alignment = center
        ws[f'B{start_row}'].font = bold

        ws.merge_cells(f'J{start_row}:L{start_row}')
        ws[f'J{start_row}'] = 'PERSONAS OCUPADAS'
        ws[f'J{start_row}'].border = border
        ws[f'J{start_row}'].alignment = center
        ws[f'J{start_row}'].font = bold

        # Fila Categorías
        row_cats = start_row + 1
        cats = ['1)SUPERVISORES O JEFES', '2)EMPLEADOS', '3)OBREROS', '4)MENORES']
        col = 2 # B
        for cat in cats:
            ws.merge_cells(start_row=row_cats, start_column=col, end_row=row_cats, end_column=col+1)
            cell = ws.cell(row_cats, col, cat)
            cell.border = border
            cell.alignment = center
            cell.font = bold
            col += 2
            
        ws.merge_cells(f'J{row_cats}:K{row_cats}')
        ws[f'J{row_cats}'] = '5)TOTALES'
        ws[f'J{row_cats}'].border = border; ws[f'J{row_cats}'].alignment = center; ws[f'J{row_cats}'].font = bold
        
        ws.merge_cells(f'L{row_cats}:L{row_cats+1}')
        ws[f'L{row_cats}'] = 'TOTAL GENERAL'
        ws[f'L{row_cats}'].border = border; ws[f'L{row_cats}'].alignment = center; ws[f'L{row_cats}'].font = bold

        # Fila Sexo
        row_sex = start_row + 2
        col = 2
        for _ in range(5): # 4 categorias + 1 total
            ws.cell(row_sex, col, 'VARONES').border = border
            ws.cell(row_sex, col).alignment = center
            ws.cell(row_sex, col+1, 'MUJERES').border = border
            ws.cell(row_sex, col+1).alignment = center
            col += 2

        # Datos
        data = self._get_data_summary(batch)
        
        rows_labels = [
            ('NUMERO', 'count'),
            ('HORAS TRABAJADAS', 'hours'),
            ('SUELDO O SALARIO', 'salary'),
            ('ENTRADAS', 'entries'),
            ('SALIDAS', 'exits')
        ]
        
        current_row = start_row + 3
        for label, key in rows_labels:
            ws.cell(current_row, 1, label).border = border
            ws.cell(current_row, 1).font = bold
            
            col = 2
            total_m = 0
            total_f = 0
            
            # Categorias
            for cat in ['supervisores', 'empleados', 'obreros', 'menores']:
                val_m = data[cat]['M'][key]
                val_f = data[cat]['F'][key]
                
                ws.cell(current_row, col, val_m).border = border
                ws.cell(current_row, col).alignment = center
                
                ws.cell(current_row, col+1, val_f).border = border
                ws.cell(current_row, col+1).alignment = center
                
                total_m += val_m
                total_f += val_f
                col += 2
            
            # Totales Sexo
            ws.cell(current_row, col, total_m).border = border
            ws.cell(current_row, col).alignment = center
            ws.cell(current_row, col+1, total_f).border = border
            ws.cell(current_row, col+1).alignment = center
            
            # Total General
            ws.cell(current_row, col+2, total_m + total_f).border = border
            ws.cell(current_row, col+2).alignment = center
            ws.cell(current_row, col+2).font = bold
            
            current_row += 1

        # Ajuste anchos
        ws.column_dimensions['A'].width = 25
        for c in range(2, 14):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 15
