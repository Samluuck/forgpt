import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from odoo import models
from odoo.modules.module import get_module_resource


class AntuxDataHubBatch(models.Model):
    _inherit = 'antux.datahub.batch'

    def generate_empleados_mensual_excel(self):
        self.ensure_one()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registro Empleados y Obreros'

        # ==================================================
        # ESTILOS
        # ==================================================
        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')

        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        gray_fill = PatternFill('solid', fgColor='D9D9D9')
        white_fill = PatternFill('solid', fgColor='FFFFFF')

        # ==================================================
        # ANCHOS DE COLUMNA
        # ==================================================
        widths = [
            5, 18, 18, 14, 14, 6, 12, 10, 6,
            12, 12, 10, 12, 12, 6, 12
        ]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # 🔴 CLAVE PARA EL LOGO
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

        # ==================================================
        # ALTURA DE FILAS
        # ==================================================
        ws.row_dimensions[1].height = 28
        for r in range(2, 6):   # A2 → A5 (logo)
            ws.row_dimensions[r].height = 35
        ws.row_dimensions[8].height = 26
        ws.row_dimensions[9].height = 26

        # ==================================================
        # BANDA SUPERIOR
        # ==================================================
        ws.merge_cells('A1:P1')
        ws['A1'] = 'Registro de Empleados y Obreros'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center
        ws['A1'].fill = gray_fill

        # ==================================================
        # LOGO MTESS ocupando A1:B2 (visual)
        # ==================================================

        # Columnas A y B
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 28

        # Filas 1 y 2
        ws.row_dimensions[1].height = 65
        ws.row_dimensions[2].height = 65

        logo_path = get_module_resource(
            'antux_data_hub', 'static', 'img', 'logo_mtess.png'
        )

        if logo_path:
            img = Image(logo_path)

            # Tamaño aproximado equivalente a A1:B2
            img.width = 450   # píxeles
            img.height = 200  # píxeles

            # Anclar en A1
            ws.add_image(img, 'A3')


        # ==================================================
        # DATOS EMPRESA
        # ==================================================
        ws.merge_cells('C2:F2')
        ws['C2'] = 'REGISTRO PATRONAL Nº:'
        ws['C2'].font = bold

        ws.merge_cells('G2:H2')
        ws['G2'] = self.company_id.vat or ''

        ws.merge_cells('I2:J2')
        ws['I2'] = 'Año'
        ws['I2'].font = bold

        ws.merge_cells('K2:L2')
        ws['K2'] = self.period_id.name or ''

        ws.merge_cells('C3:D3')
        ws['C3'] = 'Localidad:'
        ws['C3'].font = bold
        ws.merge_cells('E3:G3')
        ws['E3'] = 'ASUNCION (DISTRITO)'

        ws.merge_cells('C4:D4')
        ws['C4'] = 'Departamento:'
        ws['C4'].font = bold
        ws.merge_cells('E4:G4')
        ws['E4'] = 'CAPITAL'

        ws.merge_cells('C5:D5')
        ws['C5'] = 'Correo Electrónico:'
        ws['C5'].font = bold
        ws.merge_cells('E5:G5')
        ws['E5'] = self.company_id.email or ''

        ws.merge_cells('I3:J3')
        ws['I3'] = 'Razón Social:'
        ws['I3'].font = bold
        ws.merge_cells('K3:O3')
        ws['K3'] = self.company_id.name or ''

        ws.merge_cells('I4:J4')
        ws['I4'] = 'RUC:'
        ws['I4'].font = bold
        ws.merge_cells('K4:O4')
        ws['K4'] = self.company_id.vat or ''

        ws.merge_cells('I5:J5')
        ws['I5'] = 'Teléfono:'
        ws['I5'].font = bold
        ws.merge_cells('K5:O5')
        ws['K5'] = self.company_id.phone or ''

        # ==================================================
        # CABECERA TABLA
        # ==================================================
        ws.merge_cells('A8:A9'); ws['A8'] = 'N°'
        ws.merge_cells('B8:B9'); ws['B8'] = 'Apellidos'
        ws.merge_cells('C8:C9'); ws['C8'] = 'Nombres'
        ws.merge_cells('D8:D9'); ws['D8'] = 'Domicilio'
        ws.merge_cells('E8:E9'); ws['E8'] = 'C.I. Nº'
        ws.merge_cells('F8:F9'); ws['F8'] = 'Nacionalidad'
        ws.merge_cells('G8:G9'); ws['G8'] = 'Edad'
        ws.merge_cells('H8:H9'); ws['H8'] = 'Fecha Nac.'
        ws.merge_cells('I8:I9'); ws['I8'] = 'Est. Civil'
        ws.merge_cells('J8:L8'); ws['J8'] = 'MENORES'
        ws['J9'] = 'Fecha Nac.'
        ws['K9'] = 'Cap. Dif.'
        ws['L9'] = 'Escol.'
        ws.merge_cells('M8:M9'); ws['M8'] = 'Cargo'
        ws.merge_cells('N8:N9'); ws['N8'] = 'Fecha Entrada'
        ws.merge_cells('O8:O9'); ws['O8'] = 'Fecha Salida'
        ws.merge_cells('P8:P9'); ws['P8'] = 'Estado'

        for row in ws.iter_rows(min_row=8, max_row=9, min_col=1, max_col=16):
            for cell in row:
                cell.font = bold
                cell.alignment = center
                cell.border = thin
                cell.fill = gray_fill

        # ==================================================
        # DATOS
        # ==================================================
        row = 10
        i = 1
        for l in self.line_ids:
            ws.append([
                i,
                l.ci_number or '',
                l.last_name or '',
                l.first_name or '',
                l.job_title or '',
                l.nationality or '',
                l.profession or '',
                l.entry_date or '',
                l.work_schedule or '',
                l.children_under_18 or 0,
                l.children_with_different_abilities or 0,
                l.children_educated or 0,
                l.exit_date or '',
                l.exit_reason or '',
                l.salary_total or l.salary_base or 0.0,
                l.state or '',
            ])

            for col in range(1, 17):
                cell = ws.cell(row=row, column=col)
                cell.border = thin
                cell.alignment = center
            row += 1
            i += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
