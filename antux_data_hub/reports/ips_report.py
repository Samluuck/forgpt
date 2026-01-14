import io
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from odoo import models
from odoo.modules.module import get_module_resource

class AntuxIPSReport(models.AbstractModel):
    _name = 'antux.ips.report'
    _description = 'Reporte IPS'

    def build_import_excel(self, batch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'IPS'

        bold = Font(bold=True)
        align_center = Alignment(horizontal='center', vertical='center')

        # ==================================================
        # ANCHO DE COLUMNAS (más gruesas)
        # ==================================================
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18

        # ==================================================
        # LOGO (MISMO ESTILO QUE MTESS / EMPLEADOS)
        # ==================================================
        logo_path = get_module_resource(
            'antux_data_hub', 'static', 'img', 'logo_mtess.png'
        )

        if logo_path:
            img = Image(logo_path)
            img.width = 250   # ancho grande
            img.height = 150  # alto grande
            ws.add_image(img, 'A1')

        # ==================================================
        # CABECERA EMPRESA
        # ==================================================
        ws.merge_cells('C1:G1')
        ws['C1'] = 'REPORTE DE IPS'
        ws['C1'].font = bold
        
        ws.merge_cells('C3:F3')
        ws['C3'] = 'REGISTRO PATRONAL Nº:'
        ws['C3'].font = bold

        ws.merge_cells('G3:H3')
        ws['G3'] = batch.company_id.vat or ''

        ws.merge_cells('I3:J3')
        ws['I3'] = 'Año'
        ws['I3'].font = bold

        ws.merge_cells('K3:L3')
        ws['K3'] = batch.period_id.name or ''

        ws.merge_cells('C4:D4')
        ws['C4'] = 'Localidad:'
        ws['C4'].font = bold
        ws.merge_cells('E4:G4')
        ws['E4'] = 'ASUNCION (DISTRITO)'

        ws.merge_cells('C5:D5')
        ws['C5'] = 'Departamento:'
        ws['C5'].font = bold
        ws.merge_cells('E5:G5')
        ws['E5'] = 'CAPITAL'

        ws.merge_cells('C6:D6')
        ws['C6'] = 'Correo Electrónico:'
        ws['C6'].font = bold
        ws.merge_cells('E6:G6')
        ws['E6'] = batch.company_id.email or ''

        ws.merge_cells('I4:J4')
        ws['I4'] = 'Razón Social:'
        ws['I4'].font = bold
        ws.merge_cells('K4:O4')
        ws['K4'] = batch.company_id.name or ''

        ws.merge_cells('I5:J5')
        ws['I5'] = 'RUC:'
        ws['I5'].font = bold
        ws.merge_cells('K5:O5')
        ws['K5'] = batch.company_id.vat or ''

        ws.merge_cells('I6:J6')
        ws['I6'] = 'Teléfono:'
        ws['I6'].font = bold
        ws.merge_cells('K6:O6')
        ws['K6'] = batch.company_id.phone or ''

        # ==================================================
        # ENCABEZADOS TABLA
        # ==================================================
        headers = [
            'Nro Patronal',
            'Nro Asegurado',
            'CI',
            'Apellidos',
            'Nombres',
            'Categoría',
            'Días Trabajados',
            'Salario',
            'Mes/Año de Pago',
            'Cod. Actividad'
        ]

        header_row = 8
        ws.append([])      
        ws.append(headers) 

        for cell in ws[header_row]:
            cell.font = bold
            cell.alignment = align_center

        ws.freeze_panes = 'A9'

        # ==================================================
        # DATOS
        # ==================================================
        for line in batch.ips_line_ids:
            ws.append([
                line.patronal_number or '',
                line.insured_number or '',
                line.ci_number or '',
                line.last_name or '',
                line.first_name or '',
                'E',
                line.days_worked or 0,
                line.salary_base or 0.0,
                line.payment_month_year or '',
                line.activity_code or '',
            ])
            # Centrar la fila recién agregada
            for cell in ws[ws.max_row]:
                cell.alignment = align_center

        # ==================================================
        # SALIDA
        # ==================================================
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream


    def build_control_excel(self, batch):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reporte IPS'

        bold = Font(bold=True)
        align_center = Alignment(horizontal='center', vertical='center')

        # ==================================================
        # ANCHO DE COLUMNAS (más gruesas)
        # ==================================================
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 18

        headers = [
            'Nro Patronal',
            'Nro Asegurado',
            'CI',
            'Apellidos',
            'Nombres',
            'Categoría',
            'Días Trabajados',
            'Salario',
            'Mes/Año de Pago',
            'Cod. Actividad'
        ]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = align_center

        ws.freeze_panes = 'A2'

        for line in batch.ips_line_ids:
            ws.append([
                line.patronal_number,
                line.insured_number,
                line.ci_number,
                line.last_name,
                line.first_name,
                'E',
                line.days_worked,
                line.salary_base,
                line.payment_month_year,
                line.activity_code,
            ])
            # Centrar la fila recién agregada
            for cell in ws[ws.max_row]:
                cell.alignment = align_center

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream
