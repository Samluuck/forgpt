from odoo import fields, models
from odoo.exceptions import UserError
import io
import openpyxl
from openpyxl.styles import Font, Alignment

class AntuxDataHubBatch(models.Model):
    _name = 'antux.datahub.batch'
    _description = 'Data Hub - Lote de Procesamiento'
    _order = 'create_date desc'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    name = fields.Char(required=True, tracking=True)
    company_id = fields.Many2one('res.company',required=True,default=lambda self: self.env.company,string='Compañía')
    period_id = fields.Many2one('antux.datahub.period',required=True,tracking=True,string='Periodo')
    state = fields.Selection([('draft', 'Borrador'),('processed', 'Procesado')], default='draft', tracking=True,string='Estado')
    line_ids = fields.One2many('antux.datahub.line','batch_id',string='Registros')
    line_count = fields.Integer(compute='_compute_line_count',store=False)
    imported = fields.Boolean(string='Planilla importada',default=False,tracking=True)
    import_date = fields.Datetime(string='Fecha de importación',readonly=True)
    employees_line_ids = fields.One2many('antux.datahub.line','batch_id',string='Empleados / Obreros',compute='_compute_employees_lines',store=False)
    ips_line_ids = fields.One2many('antux.datahub.line','batch_id',string='IPS',compute='_compute_ips_lines',store=False)
    vacaciones_line_ids = fields.One2many(
        'antux.datahub.line',
        'batch_id',
        compute='_compute_vacaciones_lines',
        store=False,
    )
        
    def _compute_vacaciones_lines(self):
        for batch in self:
            batch.vacaciones_line_ids = batch.line_ids.filtered(
                lambda l:
                    (l.dias_vacaciones and l.dias_vacaciones > 0)
                    or l.desde_vacaciones
                    or l.hasta_vacaciones
            )

    
    def _compute_employees_lines(self):
        for batch in self:
            batch.employees_line_ids = batch.line_ids.filtered(
                lambda l: l.job_title
            )

    def _compute_ips_lines(self):
        for batch in self:
            batch.ips_line_ids = batch.line_ids.filtered(
                lambda l: l.patronal_number and l.insured_number
            )

    def _compute_line_count(self):
        for rec in self:
            rec.line_count = len(rec.line_ids)

    # -----------------------------------------------------
    # GENERADOR CENTRAL
    # -----------------------------------------------------
    def action_generate_by_type(self, generation_type):
        self.ensure_one()

        if not self.line_ids:
            raise UserError('El lote no tiene registros importados.')

        if generation_type == 'ips':
            return self.action_generate_ips()

        if generation_type == 'empleados_obreros':
            return self.action_generate_empleados_obreros()


        if generation_type == 'nomina':
            return self.action_generate_nomina()

        if generation_type == 'all':
            return self.action_generate_all()

        raise UserError('Tipo de generación no soportado.')

    # -----------------------------------------------------
    # IPS
    # -----------------------------------------------------
    def action_generate_ips(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/antux_datahub/ips/{self.id}',
            'target': 'self',
        }
    def generate_ips_anual_excel(self):
        self.ensure_one()
        stream = self.env['antux.ips.report'].build_control_excel(self)
        return stream, f'IPS_ANUAL_{self.name}.xlsx'    
    # -----------------------------------------------------
    # STUBS (por ahora)
    # -----------------------------------------------------
    def action_generate_mtess(self):
        raise UserError('MTESS aún no implementado.')

    def action_generate_nomina(self):
        raise UserError('Nómina aún no implementada.')

    def action_generate_all(self):
        raise UserError('Generación múltiple aún no implementada.')

    def action_generate_empleados_obreros(self):
        self.ensure_one()

        if not self.line_ids:
            raise UserError('No hay registros para generar la planilla.')

        return {
            'type': 'ir.actions.act_url',
            'url': f'/antux_datahub/empleados_obreros/{self.id}',
            'target': 'self',
        }
    def action_open_import_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Importar Planilla General',
            'res_model': 'antux.datahub.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }
    def action_open_empleados_report_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Reporte Empleados / Obreros',
            'res_model': 'antux.datahub.empleados.report.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }


    # -----------------------------------------------------
    # REPORTES IPS
    # -----------------------------------------------------
    def action_open_ips_report_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Reporte IPS',
            'res_model': 'antux.datahub.ips.report.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }


    def generate_empleados_anual_excel(self):
        self.ensure_one()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Empleados y Obreros'

        # ----------------------------
        # Encabezados
        # ----------------------------
        headers = [
            'Cédula',
            'Apellidos',
            'Nombres',
            'Cargo',
            'Nacionalidad',
            'Profesión',
            'Fecha de Entrada',
            'Horario de Trabajo',
            'Hijos Menores',
            'Menores con Capacidad Diferente',
            'Menores Escolarizados',
            'Fecha de Salida',
            'Motivo de Salida',
            'Salario',
            'Estado',
        ]

        ws.append(headers)

        bold = Font(bold=True)
        align_center = Alignment(horizontal='center', vertical='center')

        for cell in ws[1]:
            cell.font = bold
            cell.alignment = align_center

        ws.freeze_panes = 'A2'

        # ----------------------------
        # Datos
        # ----------------------------
        for l in self.line_ids:
            ws.append([
                l.ci_number or '',
                l.last_name or '',
                l.first_name or '',
                l.job_title or '',
                l.nationality or '',
                l.profession or '',
                l.entry_date or '',
                l.work_schedule or '',
                l.children_under_18 or 0,
                l.children_with_different_abilities or 0,
                l.children_educated or 0,
                l.exit_date or '',
                l.exit_reason or '',
                l.salary_total or l.salary_base or 0.0,
                l.state or '',
            ])
            # Centrar la fila recién agregada
            for cell in ws[ws.max_row]:
                cell.alignment = align_center

        # ----------------------------
        # Salida
        # ----------------------------
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream   

    def generate_ips_import_excel(self):
        stream = self.env['antux.ips.report'].build_import_excel(self)
        return stream, f'IPS_IMPORT_{self.name}.xlsx'


    # -----------------------------------------------------
    # REPORTES SUELDOS Y JORNALES
    # -----------------------------------------------------
    def action_open_sueldos_jornales_report_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Reporte Sueldos y Jornales',
            'res_model': 'antux.datahub.sueldos.jornales.report.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }

    def generate_sueldos_jornales_mensual_excel(self):
        self.ensure_one()
        return self.env['antux.sueldos.jornales.report'].build_mensual_excel(self)

    def generate_sueldos_jornales_anual_excel(self):
        self.ensure_one()
        return self.env['antux.sueldos.jornales.report'].build_control_excel(self)



    # -----------------------------------
    # VACACIONES
    # -----------------------------------

    def generate_vacaciones_mensual_excel(self):
        stream = self.env['antux.vacaciones.report'].build_mensual_excel(self)
        return stream, f'Registro_de_Vacaciones_{self.name}.xlsx'

    def generate_vacaciones_anual_excel(self):
        stream = self.env['antux.vacaciones.report'].build_anual_excel(self)
        return stream, f'VACACIONES_ANUAL_{self.name}.xlsx'


    def action_open_vacaciones_report_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Reporte Vacaciones',
            'res_model': 'antux.datahub.vacaciones.report.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }

    # -----------------------------------
    # RESUMEN GENERAL
    # -----------------------------------
    def action_open_resumen_report_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Resumen General de Personas Ocupadas',
            'res_model': 'antux.datahub.resumen.report.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }

    def action_open_salary_receipt_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Recibo de Salario',
            'res_model': 'antux.datahub.salary.receipt.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'active_id': self.id,
                'active_model': self._name,
            },
        }

    def generate_resumen_general_excel(self, report_format):
        self.ensure_one()
        if report_format == 'control':
            return self.env['antux.resumen.general.report'].build_control_excel(self)
        else:
            return self.env['antux.resumen.general.report'].build_tabular_excel(self)