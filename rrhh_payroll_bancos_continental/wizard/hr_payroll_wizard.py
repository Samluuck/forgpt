import logging
import base64
import tempfile as tmp
from datetime import datetime
from time import mktime
from openpyxl import Workbook
from odoo import api, fields, models, _
from openpyxl.styles import Alignment

logger = logging.getLogger(__name__)


class HrPayRollTxtWizard(models.TransientModel):
    _name = "hr_payroll_wizard"
    name = fields.Char('Name')
    state = fields.Selection([
        ('view', 'View'),
        ('get', 'Get')
    ], string='State', required=True, default='view')
    data = fields.Binary('File', readonly=True)

    def get_file_txt(self):
        print("########################   def get_file_txt(self):    ##################################")
        payroll_run = self.env['hr.payslip.run'].browse(self.env.context.get('active_id'))
        buf_file = None
        if payroll_run:
            print("######################  if payroll_run:   ##########################################")
            dt = datetime.now()
            tm = mktime(dt.timetuple())
            file_name = '{0}_{1}.txt'.format(payroll_run.id, tm)
            file_with_path = '{0}/{1}'.format(tmp.gettempdir(), file_name)

            payslips = self.env['hr.payslip'].search([('payslip_run_id', '=', payroll_run.id)])
            company = self.env.user.company_id  # Obtener la empresa actual de la sesión del usuario
            nro_cuenta = company.nro_cuenta if company.nro_cuenta else 'N/A'

            records = []
            for payslip in payslips:
                employee = payslip.employee_id
                bank_account = nro_cuenta
                print(bank_account)

                # Filtrar reglas salariales
                salary_lines = payslip.line_ids
                net_lines = salary_lines.filtered(lambda l: l.code in ['NET', 'NETO'])
                ant_lines = salary_lines.filtered(lambda l: l.code == 'ANT')

                if net_lines:
                    salary = sum(net_lines.mapped('amount'))  # Si hay NET/NETO, solo sumamos estas
                elif ant_lines:
                    salary = sum(ant_lines.mapped('amount'))  # Si no hay NET/NETO, tomamos ANT
                else:
                    salary = 0  # En caso de que no haya ninguna de las reglas

                aguinaldo = 'SI' if payslip.struct_id.es_aguinaldo else 'NO'
                record = (
                    employee.identification_id or 'N/A',
                    bank_account,
                    f"Nomina de {payroll_run.date_end.strftime('%B %Y')}",
                    int(salary),
                    aguinaldo,
                )
                records.append(record)

            with open(file_with_path, "w") as f:
                if records:
                    for line in records:
                        txt_value = "{0} , {1} , {2}  , {3} , {4}".format(*line)
                        f.write("{0}\n".format(txt_value.strip()))
                        print(*line)
                else:
                    f.write("No se encontraron registros.")

            with open(file_with_path, "rb") as f:
                buf_file = base64.b64encode(f.read())
            name = "Archivo_Banco.txt"
            self.write({'state': 'get', 'data': buf_file, 'name': name})

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'hr_payroll_wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'res_id': self.id,
                'views': [(False, 'form')],
                'target': 'new',
            }

    def get_file_excel(self):
        payroll_run = self.env['hr.payslip.run'].browse(self.env.context.get('active_id'))
        if payroll_run:
            # Crear un archivo de Excel
            wb = Workbook()
            ws = wb.active

            # Agregar encabezados
            headers = ['Numero de CI', 'Cuenta Debito', 'Concepto', 'Importe', 'Aguinaldo']
            ws.append(headers)

            # Estilo para centrar el contenido de las celdas
            centered_alignment = Alignment(horizontal='center', vertical='center')

            # Centrar encabezados
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                  cell.alignment = centered_alignment


            # Función para centrar el contenido de una fila
            def center_row(row_idx, col_start, col_end):
                for cell in ws[row_idx:row_idx]:
                    for c in cell:
                        c.alignment = Alignment(horizontal="center")

            # Obtener las nóminas relacionadas con el payroll_run
            payslips = self.env['hr.payslip'].search([
                ('payslip_run_id', '=', payroll_run.id)
            ])
            company = self.env.user.company_id  # Obtener la empresa actual de la sesión del usuario
            nro_cuenta = company.nro_cuenta if company.nro_cuenta else 'N/A'
            for payslip in payslips:
                employee = payslip.employee_id
                ci = employee.identification_id
                bank_account = nro_cuenta    #if employee.bank_account_id else 'N/A'
                salary_line = payslip.line_ids.filtered(lambda l: l.code == 'NETO' or l.code == 'NET')
                salary = salary_line.amount if salary_line.amount else 0

                # Aquí obtenemos el concepto de "Nomina de Month YYYY"
                concept = 'Nomina de ' + payroll_run.date_end.strftime('%B %Y')

                # Aguinaldo (valor estático 'NO' por ahora)
                aguinaldo = 'SI' if payslip.struct_id.es_aguinaldo else 'NO'

                # Agregar los datos a la hoja de Excel
                data_row = [str(ci), str(bank_account), concept, salary, aguinaldo]
                ws.append(data_row)

                # Establecer el formato de la celda para ci y cuenta como texto
                ws.cell(row=ws.max_row, column=1).number_format = '@'  # Número de CI
                ws.cell(row=ws.max_row, column=2).number_format = '@'  # Cuenta Débito

                # Centrar el contenido de la fila actual
                for cell in ws[ws.max_row]:
                    cell.alignment = centered_alignment

            # Crear un archivo temporal de Excel
            excel_file = tmp.NamedTemporaryFile(suffix='.xlsx', delete=False)
            wb.save(excel_file.name)
            excel_file.close()

            # Leer el archivo temporal y codificarlo en base64
            with open(excel_file.name, "rb") as f:
                buf_file = base64.b64encode(f.read())

            # Definir el nombre del archivo
            name = "Archivo_Banco.xlsx"

            # Actualizar el estado y los datos del archivo en el modelo
            self.write({'state': 'get', 'data': buf_file, 'name': name})

            return {
                'type': 'ir.actions.act_window',
                'res_model': 'hr_payroll_wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'res_id': self.id,
                'views': [(False, 'form')],
                'target': 'new',
            }